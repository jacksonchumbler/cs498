import pandas as pd
import sys
from openpyxl import *
from tkinter import *


user_df = pd.read_csv('users.csv')
form_df = pd.read_csv('forms.csv')
current_user = "invalid_user"


def main():

    login()
    menu()

    print(user_df)
    print(form_df)
    user_df.to_csv('users.csv', index = False)
    form_df.to_csv('forms.csv', index = False)

def menu():
    global current_user
    while True:
        print("\nWelcome, {}, to the Coroner Database".format(user_df.loc[user_df['email'] == current_user, 'name'].item()))
        print("Available Operations:")
        print("Create New Form [0]")
        print("View Form       [1]")
        print("Secure Logout   [2]")

        choice = int(input('Enter a choice: '))
        print()
        if   choice == 0:
            # form()
            gui_form()
        elif choice == 1:
            view()
        elif choice == 2:
            # Securely clears login
            current_user = ""
            break
        else:
            print("Invalid Input!")

def form():
    print("Please provide the following information...")
    name_d = input("Name of Deceased: ")
    place_d = input("Address of Death: ")
    time_d = input("Death Time: ")
    form_df.loc[len(form_df)] = [current_user, name_d, place_d, time_d]
    print()

def view():
    # print("Under Construction!")
    # print(form_df.loc[form_df['author'] == current_user])
    for index, row in form_df.loc[form_df['author'] == current_user].iterrows():
        print(row)

    return

def logout():
    pass

def login():
    attempts = 0
    user_type = input('Are you a (new/existing) user?\n')
    if user_type == 'new':
        generate_user()
        print("Successfully Added Account")
    while attempts < 3:
        print("\nPlease Login...")
        email = input('Enter login email: ')
        passwd = input('Enter Password: ')

        if email in user_df['email'].values:
            login_user = user_df.loc[user_df['email'] == email]
            if passwd in login_user['passwd'].values:
                global current_user
                current_user = email
                return
        print("\nIncorrect Login")
        attempts += 1
    print("\nExceeded Login Retries")
    sys.exit()

def generate_user():
    name = input('Enter your name: ')
    email = input('Enter your email: ')
    passwd  = input('Enter your password: ')
    user_df.loc[len(user_df)] = [name, email, passwd]
    return

# import openpyxl and tkinter modules

# globally declare wb and sheet variable

# opening the existing excel file
# wb = load_workbook('C:\\Users\\Admin\\Desktop\\excel.xlsx')

# create the sheet object
# sheet = wb.active

'''
def excel():
	
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Course"
    sheet.cell(row=1, column=3).value = "Semester"
    sheet.cell(row=1, column=4).value = "Form Number"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"
'''

# Function to set focus (cursor)
def focus1(event):
    # set focus on the course_field box
    first_name_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the sem_field box
    last_name_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the form_no_field box
    address_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the contact_no_field box
    city_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the email_id_field box
    county_field.focus_set()


# Function to set focus
def focus6(event):
    # set focus on the address_field box
    state_field.focus_set()

def focus7(event):
    zip_code_field.focus_set()
def focus8(event):
    ssn_field.focus_set()
def focus9(event):
    age_field.focus_set()
def focus10(event):
    cause_death_field.focus_set()

# Function for clearing the
# contents of text entry boxes
def clear():
	
    # clear the content of text entry box
    first_name_field.delete(0, END)
    last_name_field.delete(0, END)
    address_field.delete(0, END)
    city_field.delete(0, END)
    county_field.delete(0, END)
    state_field.delete(0, END)
    zip_code_field.delete(0, END)
    ssn_field.delete(0, END)
    age_field.delete(0, END)
    cause_death_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert(first_name_field, last_name_field, address_field, city_field,
           county_field, state_field, zip_code_field, ssn_field, age_field,
           cause_death_field):
    # if user not fill any entry
    # then print "empty input"
    if (first_name_field.get() == "" and
        last_name_field.get() == "" and
        address_field.get() == "" and
        city_field.get() == "" and
        county_field.get() == "" and
        state_field.get() == "" and
        zip_code_field.get() == "" and
        ssn_field.get() == "" and
        age_field.get() == "" and
        cause_death.get() == ""):
        print("empty input")

    else:
        '''
        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        # save the file
        # wb.save('C:\\Users\\Admin\\Desktop\\excel.xlsx')
        wb.save('./excel.xlsx')
        '''
        form_df.loc[len(form_df)] = [current_user, first_name_field.get(), last_name_field.get(),
                                     address_field.get(), city_field.get(),
                           county_field.get(), state_field.get(), zip_code_field.get(),
                                     ssn_field.get(), age_field.get(),
                           cause_death_field.get()] # [current_user, name_d, place_d, time_d]

        # set focus on the name_field box
        first_name_field.focus_set()

        # call the clear() function
        # clear()


# Driver code
# if __name__ == "__main__":
def gui_form():
	
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light blue')

    # set the title of GUI window
    root.title("registration form")

    # set the configuration of GUI window
    root.geometry("500x300")

    # excel()

    # create a Form label
    heading = Label(root, text="Form", bg="light blue")

    # create a Name label
    first_name = Label(root, text="First Name", bg="light blue")
    last_name = Label(root, text="Last Name", bg="light blue")
    address = Label(root, text="Address", bg="light blue")
    city = Label(root, text="City", bg="light blue")
    county = Label(root, text="County of Residence", bg="light blue")
    state = Label(root, text="State of Residence", bg="light blue")
    zip_code = Label(root, text="Zip Code", bg="light blue")
    ssn = Label(root, text="SSN#", bg="light blue")
    age = Label(root, text="Age", bg="light blue")
    cause_death = Label(root, text="Cause of Death", bg="light blue")

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    first_name.grid(row=0, column=0)
    last_name.grid(row=0, column=2)
    address.grid(row=1, column=0)
    city.grid(row=1, column=2)
    county.grid(row=1, column=4)
    state.grid(row=2, column=0)
    zip_code.grid(row=2, column=2)
    ssn.grid(row=2, column=4)
    age.grid(row=3, column=0)
    cause_death.grid(row=4, column=0)

    # create a text entry box
    # for typing the information

    first_name_field = Entry(root)
    last_name_field = Entry(root)
    address_field = Entry(root)
    city_field = Entry(root)
    county_field = Entry(root)
    state_field = Entry(root)
    zip_code_field = Entry(root)
    ssn_field = Entry(root)
    age_field = Entry(root)
    cause_death_field = Entry(root)

    first_name_field.bind("<Return>", focus1)
    last_name_field.bind("<Return>", focus2)
    address_field.bind("<Return>", focus3)
    city_field.bind("<Return>", focus4)
    county_field.bind("<Return>", focus5)
    state_field.bind("<Return>", focus6)
    zip_code_field.bind("<Return>", focus7)
    ssn_field.bind("<Return>", focus8)
    age_field.bind("<Return>", focus9)
    cause_death_field.bind("<Return>", focus10)

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    first_name_field.grid(row=0, column=1, ipadx="100")
    last_name_field.grid(row=0, column=3, ipadx="100")
    address_field.grid(row=1, column=1, ipadx="100")
    city_field.grid(row=1, column=3, ipadx="100")
    county_field.grid(row=1, column=5, ipadx="100")
    state_field.grid(row=2, column=1, ipadx="100")
    zip_code_field.grid(row=2, column=3, ipadx="100")
    ssn_field.grid(row=2, column=5, ipadx="100")
    age_field.grid(row=3, column=1, ipadx="100")
    cause_death_field.grid(row=4, column=1, ipadx="100")

    # call excel function
    # excel()

    # create a Submit Button and place into the root window
    # submit = Button(root, text="Submit", fg="Black", bg="Red", command=insert)
    submit = Button(root, text="Submit", fg="Black", bg="Red", command= lambda:
                    [insert(first_name_field, last_name_field, address_field, city_field,
                           county_field, state_field, zip_code_field, ssn_field, age_field,
                           cause_death_field), root.destroy()])
    submit.grid(row=11, column=1)
    # start the GUI
    root.mainloop()





if __name__ == '__main__':
    main()
